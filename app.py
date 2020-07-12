import time
import requests
import json
import openpyxl

wb = openpyxl.Workbook()

xlpath = 'excel/data.xlsx'
wb.create_sheet(index = 0 , title = "sheet1")
wb.create_sheet(index = 1 , title = "sheet2")
sheet1 = wb['sheet1']
sheet2 = wb['sheet2']


def cityCleaner(data):
    li = []
    for i in data:
        if i['country']=='IN':
            li.append([i['id'],i['name'],i['country']])
    return li

def writexl(li):
    #writing sheet1 headers
    sheet1.cell(row=1, column=1).value = 'City Name'
    sheet1.cell(row=1, column=2).value = 'Temperature(F)'
    sheet1.cell(row=1, column=3).value = 'Temperature(C)'
    sheet1.cell(row=1,column=4).value = 'UpdateStatus'

    #writing sheet 2
    sheet2.cell(row=1, column=1).value = 'City ID'
    sheet2.cell(row=1, column=2).value = 'City Name'
    sheet2.cell(row=1, column=3).value = 'Country Code'
    for i in range(2,len(li)):
        sheet2.cell(row=i, column=1).value = li[i][0]
        sheet2.cell(row=i, column=2).value = li[i][1]
        sheet2.cell(row=i, column=3).value = li[i][2]
    wb.save(xlpath)

def cityNames():
    f = open('Data/city.json')
    data = json.load(f)
    li = cityCleaner(data)
    writexl(li)
    return li

def clean(data):
    result = []
    K = data['main']['temp']
    c = round(K - 273.15,2)
    f = round((c*(9/5))+32,2)
    result.append([c,f])
    result.append(data['name'])
    return result


def txt(cityData):
    f = open('Data/data.txt','w')
    for data in cityData:
        f.write(f'{data[1]},{data[0][1]}\n')
    f.close


def weather():
    cityData = []
    cities = ['faridabad','delhi','kochi','mumbai','patna']
    for city in cities:
        response = requests.get(f'https://api.openweathermap.org/data/2.5/weather?q={city}&appid=f3232d76c8092c56b0f75fb1dbe9ffa9')
        data = response.json()
        cityData.append(clean(data))
    txt(cityData)

#Running File

run = True
cityNames()

print('Stating System .... To Stop Code Press Ctrl+C!!!')
while run:
    try:
        weather()
        print('running..')
        time.sleep(3)
    except:
        print('\t\nShutting Down')
        time.sleep(1)
        run = False